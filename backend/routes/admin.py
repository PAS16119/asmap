import base64
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity, get_jwt
from functools import wraps
from models import (
    db, User, Teacher, Subject, Class, Student,
    CoordinatorClassAssignment, Payment, MonitoringRecord,
    AcademicYear, SchoolSettings
)
from sqlalchemy import func
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import datetime

admin_bp = Blueprint('admin', __name__)


# â”€â”€ Decorators â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def admin_required(fn):
    @wraps(fn)
    @jwt_required()
    def wrapper(*args, **kwargs):
        if get_jwt().get('role') != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        return fn(*args, **kwargs)
    return wrapper


def _active_year_id():
    y = AcademicYear.query.filter_by(is_active=True).first()
    return y.id if y else None


def _active_year():
    return AcademicYear.query.filter_by(is_active=True).first()


# â”€â”€ SCHOOL SETTINGS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin_bp.route('/settings/public', methods=['GET'])
def settings_public():
    """Unauthenticated â€” used by login page."""
    s = SchoolSettings.query.first()
    if not s:
        return jsonify({'school_name': 'Your School', 'school_short': '', 'school_motto': '', 'school_logo': ''})
    return jsonify(s.to_public_dict())


@admin_bp.route('/settings', methods=['GET'])
@jwt_required()
def get_settings():
    s = SchoolSettings.query.first()
    if not s:
        s = SchoolSettings()
        db.session.add(s)
        db.session.commit()
    return jsonify(s.to_dict())


@admin_bp.route('/settings', methods=['PUT'])
@admin_required
def update_settings():
    data = request.get_json(silent=True) or {}
    s = SchoolSettings.query.first()
    if not s:
        s = SchoolSettings()
        db.session.add(s)

    if 'school_name'    in data: s.school_name    = data['school_name'].strip()
    if 'school_short'   in data: s.school_short   = data['school_short'].strip()
    if 'school_motto'   in data: s.school_motto   = data['school_motto'].strip()
    if 'normal_start'   in data: s.normal_start   = data['normal_start'].strip()
    if 'extended_start' in data: s.extended_start = data['extended_start'].strip()

    db.session.commit()
    return jsonify(s.to_dict())


@admin_bp.route('/settings/logo', methods=['POST'])
@admin_required
def upload_logo():
    if 'logo' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['logo']
    if file.filename == '':
        return jsonify({'error': 'Empty filename'}), 400

    allowed = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
    ext = file.filename.rsplit('.', 1)[-1].lower()
    if ext not in allowed:
        return jsonify({'error': 'Invalid file type. Use PNG, JPG or GIF'}), 400

    img_bytes = file.read()
    if len(img_bytes) > 2 * 1024 * 1024:
        return jsonify({'error': 'Logo must be under 2 MB'}), 400

    mime_map = {'jpg': 'image/jpeg', 'jpeg': 'image/jpeg', 'png': 'image/png',
                'gif': 'image/gif', 'webp': 'image/webp'}
    mime = mime_map.get(ext, 'image/png')
    b64  = base64.b64encode(img_bytes).decode('utf-8')
    data_url = f'data:{mime};base64,{b64}'

    s = SchoolSettings.query.first()
    if not s:
        s = SchoolSettings()
        db.session.add(s)
    s.school_logo = data_url
    db.session.commit()
    return jsonify({'message': 'Logo uploaded', 'logo': data_url})


# â”€â”€ ACADEMIC YEARS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin_bp.route('/years', methods=['GET'])
@jwt_required()
def list_years():
    years = AcademicYear.query.order_by(AcademicYear.created_at.desc()).all()
    return jsonify([y.to_dict() for y in years])


@admin_bp.route('/years/active', methods=['GET'])
@jwt_required()
def get_active_year():
    y = _active_year()
    return jsonify(y.to_dict() if y else None)


@admin_bp.route('/years', methods=['POST'])
@admin_required
def create_year():
    data = request.get_json(silent=True) or {}
    label = (data.get('label') or '').strip()
    if not label:
        return jsonify({'error': 'Year label required (e.g. 2025/2026)'}), 400
    if AcademicYear.query.filter_by(label=label).first():
        return jsonify({'error': 'Year already exists'}), 409
    y = AcademicYear(label=label)
    db.session.add(y)
    db.session.commit()
    return jsonify(y.to_dict()), 201


@admin_bp.route('/years/<int:yid>/activate', methods=['POST'])
@admin_required
def activate_year(yid):
    AcademicYear.query.update({'is_active': False})
    y = AcademicYear.query.get_or_404(yid)
    y.is_active = True
    db.session.commit()
    return jsonify(y.to_dict())


# â”€â”€ DASHBOARD â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin_bp.route('/dashboard', methods=['GET'])
@admin_required
def dashboard():
    year_id = request.args.get('year_id') or _active_year_id()

    p_query = Payment.query
    s_query = Student.query.filter_by(is_active=True)
    m_query = MonitoringRecord.query
    if year_id:
        p_query = p_query.filter_by(year_id=year_id)
        s_query = s_query.filter_by(year_id=year_id)
        m_query = m_query.filter_by(year_id=year_id)

    total_collected = db.session.query(func.sum(Payment.amount))
    if year_id:
        total_collected = total_collected.filter(Payment.student.has(year_id=year_id) == year_id)
    total_collected = float(total_collected.scalar() or 0)

    # Purpose breakdown
    pq = db.session.query(Payment.purpose, func.sum(Payment.amount).label('t'), func.count(Payment.id).label('c'))
    if year_id:
        pq = pq.filter(Payment.student.has(year_id=year_id) == year_id)
    purpose_rows = pq.group_by(Payment.purpose).all()

    # Class breakdown
    cq = (db.session.query(Class.name, func.sum(Payment.amount).label('t'), func.count(Payment.id).label('c'))
          .join(Student, Student.id == Payment.student_id)
          .join(Class, Class.id == Student.class_id))
    if year_id:
        cq = cq.filter(Payment.student.has(year_id=year_id) == year_id)
    class_rows = cq.group_by(Class.name).order_by(Class.name).all()

    return jsonify({
        'total_students':      s_query.count(),
        'total_teachers':      Teacher.query.filter_by(is_active=True).count(),
        'total_classes':       Class.query.count(),
        'total_users':         User.query.filter_by(is_active=True).count(),
        'total_collected':     total_collected,
        'unconfirmed_payments':p_query.filter(Payment.is_confirmed == False).count(),
        'total_monitoring':    m_query.count(),
        'payment_count':       p_query.count(),
        'purpose_breakdown':   [{'purpose': r.purpose, 'total': float(r.t), 'count': r.c} for r in purpose_rows],
        'class_breakdown':     [{'class_name': r.name, 'total': float(r.t), 'count': r.c} for r in class_rows],
    })


# â”€â”€ USERS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin_bp.route('/users', methods=['GET'])
@admin_required
def list_users():
    return jsonify([u.to_dict() for u in User.query.order_by(User.created_at.desc()).all()])


@admin_bp.route('/users', methods=['POST'])
@admin_required
def create_user():
    data = request.get_json(silent=True) or {}
    for f in ['username', 'password', 'role']:
        if not data.get(f):
            return jsonify({'error': f'{f} is required'}), 400
    if data['role'] not in ('admin', 'coordinator', 'supervisor'):
        return jsonify({'error': 'Invalid role'}), 400
    if User.query.filter_by(username=data['username'].strip()).first():
        return jsonify({'error': 'Username already exists'}), 409
    if len(data['password']) < 6:
        return jsonify({'error': 'Password must be at least 6 characters'}), 400

    u = User(
        username=data['username'].strip(),
        email=(data.get('email') or '').strip() or None,
        role=data['role'],
        full_name=(data.get('full_name') or '').strip() or None,
        created_by=get_jwt_identity()
    )
    u.set_password(data['password'])
    db.session.add(u)
    db.session.commit()
    return jsonify(u.to_dict()), 201


@admin_bp.route('/users/<int:uid>', methods=['PUT'])
@admin_required
def update_user(uid):
    u = User.query.get_or_404(uid)
    data = request.get_json(silent=True) or {}
    if 'full_name' in data:  u.full_name = data['full_name']
    if 'email'     in data:  u.email     = data['email'] or None
    if 'is_active' in data:  u.is_active = data['is_active']
    if 'role'      in data and data['role'] in ('admin', 'coordinator', 'supervisor'):
        u.role = data['role']
    if data.get('new_password'):
        if len(data['new_password']) < 6:
            return jsonify({'error': 'Password must be at least 6 characters'}), 400
        u.set_password(data['new_password'])
    db.session.commit()
    return jsonify(u.to_dict())


# â”€â”€ TEACHERS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin_bp.route('/teachers', methods=['GET'])
@jwt_required()
def list_teachers():
    teachers = Teacher.query.filter_by(is_active=True).order_by(Teacher.full_name).all()
    return jsonify([t.to_dict() for t in teachers])


@admin_bp.route('/teachers', methods=['POST'])
@admin_required
def create_teacher():
    data = request.get_json(silent=True) or {}
    name = (data.get('full_name') or '').strip()
    if not name:
        return jsonify({'error': 'Teacher name required'}), 400
    t = Teacher(
        full_name=name,
        subject=(data.get('subject') or '').strip() or None,
        created_by=get_jwt_identity()
    )
    db.session.add(t)
    db.session.commit()
    return jsonify(t.to_dict()), 201


@admin_bp.route('/teachers/<int:tid>', methods=['PUT'])
@admin_required
def update_teacher(tid):
    t = Teacher.query.get_or_404(tid)
    data = request.get_json(silent=True) or {}
    if 'full_name' in data: t.full_name = data['full_name'].strip()
    if 'subject'      in data: t.subject      = data['subject'].strip() or None
    if 'is_active'    in data: t.is_active    = data['is_active']
    db.session.commit()
    return jsonify(t.to_dict())


@admin_bp.route('/teachers/<int:tid>', methods=['DELETE'])
@admin_required
def delete_teacher(tid):
    t = Teacher.query.get_or_404(tid)
    t.is_active = False
    db.session.commit()
    return jsonify({'message': 'Teacher deactivated'})


# â”€â”€ SUBJECTS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin_bp.route('/subjects', methods=['GET'])
@jwt_required()
def list_subjects():
    return jsonify([s.to_dict() for s in Subject.query.order_by(Subject.name).all()])


@admin_bp.route('/subjects', methods=['POST'])
@admin_required
def create_subject():
    data = request.get_json(silent=True) or {}
    name = (data.get('subject_name') or '').strip()
    if not name:
        return jsonify({'error': 'Subject name required'}), 400
    if Subject.query.filter_by(name=name).first():
        return jsonify({'error': 'Subject already exists'}), 409
    s = Subject(name=name)
    db.session.add(s)
    db.session.commit()
    return jsonify(s.to_dict()), 201


@admin_bp.route('/subjects/<int:sid>', methods=['DELETE'])
@admin_required
def delete_subject(sid):
    s = Subject.query.get_or_404(sid)
    db.session.delete(s)
    db.session.commit()
    return jsonify({'message': 'Deleted'})


# â”€â”€ CLASSES â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin_bp.route('/classes', methods=['GET'])
@jwt_required()
def list_classes():
    return jsonify([c.to_dict() for c in Class.query.order_by(Class.name).all()])


@admin_bp.route('/classes', methods=['POST'])
@admin_required
def create_class():
    data = request.get_json(silent=True) or {}
    name = (data.get('class_name') or '').strip()
    if not name:
        return jsonify({'error': 'Class name required'}), 400
    if Class.query.filter_by(name=name).first():
        return jsonify({'error': 'Class already exists'}), 409
    c = Class(name=name)
    db.session.add(c)
    db.session.commit()
    return jsonify(c.to_dict()), 201


@admin_bp.route('/classes/<int:cid>', methods=['DELETE'])
@admin_required
def delete_class(cid):
    c = Class.query.get_or_404(cid)
    db.session.delete(c)
    db.session.commit()
    return jsonify({'message': 'Deleted'})


# â”€â”€ STUDENTS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin_bp.route('/students', methods=['GET'])
@jwt_required()
def list_students():
    year_id  = request.args.get('year_id') or _active_year_id()
    class_id = request.args.get('class_id')
    q = Student.query.filter_by(is_active=True)
    if year_id:  q = q.filter_by(year_id=int(year_id))
    if class_id: q = q.filter_by(class_id=int(class_id))
    return jsonify([s.to_dict() for s in q.order_by(Student.name).all()])


@admin_bp.route('/students', methods=['POST'])
@admin_required
def create_student():
    data = request.get_json(silent=True) or {}
    name     = (data.get('name') or '').strip()
    class_id = data.get('class_id')
    if not name or not class_id:
        return jsonify({'error': 'Student name and class are required'}), 400
    year_id = data.get('year_id') or _active_year_id()
    s = Student(name=name, class_id=int(class_id),
                year_id=year_id)
    db.session.add(s)
    db.session.commit()
    return jsonify(s.to_dict()), 201


@admin_bp.route('/students/<int:sid>', methods=['PUT'])
@admin_required
def update_student(sid):
    s = Student.query.get_or_404(sid)
    data = request.get_json(silent=True) or {}
    if 'name' in data: s.name = data['student_name'].strip()
    if 'class_id'     in data: s.class_id     = int(data['class_id'])
    if 'is_active'    in data: s.is_active    = data['is_active']
    db.session.commit()
    return jsonify(s.to_dict())


@admin_bp.route('/students/import', methods=['POST'])
@admin_required
def import_students():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Only Excel files (.xlsx/.xls) accepted'}), 400

    year_id    = request.form.get('year_id') or _active_year_id()
    creator_id = get_jwt_identity()
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    created, skipped, errors = 0, 0, []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        sname = str(row[0]).strip()
        cname = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        if not sname or not cname:
            errors.append(f'Row {i}: missing name or class')
            continue

        cls = Class.query.filter_by(name=cname).first()
        if not cls:
            cls = Class(name=cname)
            db.session.add(cls)
            db.session.flush()

        if Student.query.filter_by(name=sname, class_id=cls.id, year_id=year_id).first():
            skipped += 1
            continue

        db.session.add(Student(name=sname, class_id=cls.id,
                               year_id=year_id))
        created += 1

    db.session.commit()
    return jsonify({'message': f'{created} imported, {skipped} skipped', 'created': created,
                    'skipped': skipped, 'errors': errors[:20]})


@admin_bp.route('/students/template', methods=['GET'])
@jwt_required()
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students Import'
    bold = Font(bold=True)
    ws.append(['Student Name', 'Class'])
    ws['A1'].font = bold; ws['B1'].font = bold
    ws.append(['John Mensah', 'SHS1A'])
    ws.append(['Abena Owusu', 'SHS2B'])
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='students_import_template.xlsx')


@admin_bp.route('/students/export', methods=['GET'])
@admin_required
def export_students():
    year_id = request.args.get('year_id') or _active_year_id()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Students'
    hfont = Font(bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='0A1628', end_color='0A1628', fill_type='solid')
    ws.append(['Student Name', 'Class', 'Academic Year'])
    for c in ws[1]: c.font = hfont; c.fill = hfill

    q = Student.query.filter_by(is_active=True)
    if year_id: q = q.filter_by(year_id=int(year_id))
    for s in q.order_by(Student.name).all():
        ws.append([s.name, s.class_.name if s.class_ else '',
                   s.year.label if s.year else "" if s.academic_year else ''])

    for col, w in [('A', 35), ('B', 20), ('C', 15)]:
        ws.column_dimensions[col].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='students_export.xlsx')


# â”€â”€ COORDINATOR ASSIGNMENTS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin_bp.route('/coordinator/assignments', methods=['GET'])
@admin_required
def list_assignments():
    return jsonify([a.to_dict() for a in CoordinatorClassAssignment.query.all()])


@admin_bp.route('/coordinator/assign', methods=['POST'])
@admin_required
def assign_coordinator():
    data = request.get_json(silent=True) or {}
    cid, clid = data.get('coordinator_id'), data.get('class_id')
    if not cid or not clid:
        return jsonify({'error': 'coordinator_id and class_id required'}), 400
    u = User.query.get(cid)
    if not u or u.role != 'coordinator':
        return jsonify({'error': 'User is not a coordinator'}), 400
    if CoordinatorClassAssignment.query.filter_by(coordinator_id=cid, class_id=clid).first():
        return jsonify({'error': 'Assignment already exists'}), 409
    a = CoordinatorClassAssignment(coordinator_id=int(cid), class_id=int(clid))
    db.session.add(a)
    db.session.commit()
    return jsonify(a.to_dict()), 201


@admin_bp.route('/coordinator/assign/<int:aid>', methods=['DELETE'])
@admin_required
def remove_assignment(aid):
    a = CoordinatorClassAssignment.query.get_or_404(aid)
    db.session.delete(a)
    db.session.commit()
    return jsonify({'message': 'Assignment removed'})


# â”€â”€ PAYMENTS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin_bp.route('/payments', methods=['GET'])
@admin_required
def list_payments():
    year_id   = request.args.get('year_id') or _active_year_id()
    class_id  = request.args.get('class_id')
    student_id= request.args.get('student_id')
    confirmed = request.args.get('confirmed')

    q = Payment.query
    if year_id:   q = q.filter_by(year_id=int(year_id))
    if class_id:  q = q.join(Student).filter(Student.class_id == int(class_id))
    if student_id:q = q.filter(Payment.student_id == int(student_id))
    if confirmed is not None and confirmed != '':
        q = q.filter(Payment.is_confirmed == (confirmed.lower() == 'true'))

    return jsonify([p.to_dict() for p in q.order_by(Payment.created_at.desc()).all()])


@admin_bp.route('/payments', methods=['POST'])
@admin_required
def record_payment():
    data = request.get_json(silent=True) or {}
    for f in ['student_id', 'amount', 'purpose']:
        if not data.get(f):
            return jsonify({'error': f'{f} is required'}), 400

    uid     = get_jwt_identity()
    year_id = data.get('year_id') or _active_year_id()
    p = Payment(
        student_id=int(data['student_id']),
        amount=float(data['amount']),
        purpose=data['purpose'],
        notes=data.get('notes'),
        collected_by_user_id=uid,
        year_id=year_id,
        is_confirmed=True,
        confirmed_by_admin_id=uid
    )
    db.session.add(p)
    db.session.commit()
    return jsonify(p.to_dict()), 201


@admin_bp.route('/payments/<int:pid>/confirm', methods=['POST'])
@admin_required
def confirm_payment(pid):
    p = Payment.query.get_or_404(pid)
    if p.is_confirmed:
        return jsonify({'error': 'Payment already confirmed'}), 400
    p.is_confirmed = True
    p.confirmed_by_admin_id = get_jwt_identity()
    db.session.commit()
    return jsonify(p.to_dict())


@admin_bp.route('/payments/<int:pid>', methods=['PUT'])
@admin_required
def update_payment(pid):
    p = Payment.query.get_or_404(pid)
    data = request.get_json(silent=True) or {}
    if 'amount'  in data: p.amount  = float(data['amount'])
    if 'purpose' in data: p.purpose = data['purpose']
    if 'notes'   in data: p.notes   = data['notes']
    db.session.commit()
    return jsonify(p.to_dict())


@admin_bp.route('/payments/<int:pid>', methods=['DELETE'])
@admin_required
def delete_payment(pid):
    p = Payment.query.get_or_404(pid)
    db.session.delete(p)
    db.session.commit()
    return jsonify({'message': 'Payment deleted'})


@admin_bp.route('/payments/student/<int:student_id>', methods=['GET'])
@jwt_required()
def student_payment_breakdown(student_id):
    year_id = request.args.get('year_id') or _active_year_id()
    q = Payment.query.filter_by(student_id=student_id)
    if year_id: q = q.filter_by(year_id=int(year_id))
    payments = q.order_by(Payment.created_at.desc()).all()
    total = sum(float(p.amount) for p in payments)
    return jsonify({'total': total, 'payment_count': len(payments),
                    'payments': [p.to_dict() for p in payments]})


# â”€â”€ PAYMENT REPORTS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin_bp.route('/reports/payments', methods=['GET'])
@admin_required
def payment_reports():
    year_id = request.args.get('year_id') or _active_year_id()

    def filtered(q):
        return q.filter(Payment.student.has(year_id=year_id) == year_id) if year_id else q

    by_class = (db.session.query(Class.id.label('cid'), Class.name,
                                  func.sum(Payment.amount).label('t'),
                                  func.count(Payment.id).label('c'))
                .join(Student, Student.id == Payment.student_id)
                .join(Class, Class.id == Student.class_id))
    by_class = filtered(by_class).group_by(Class.id, Class.name).order_by(Class.name).all()

    by_purpose = db.session.query(Payment.purpose,
                                   func.sum(Payment.amount).label('t'),
                                   func.count(Payment.id).label('c'))
    by_purpose = filtered(by_purpose).group_by(Payment.purpose).order_by(func.sum(Payment.amount).desc()).all()

    by_collector = (db.session.query(User.id, User.full_name, User.username,
                                      func.sum(Payment.amount).label('t'),
                                      func.count(Payment.id).label('c'))
                    .join(User, User.id == Payment.collected_by_user_id))
    by_collector = filtered(by_collector).group_by(User.id).order_by(func.sum(Payment.amount).desc()).all()

    grand = sum(float(r.t) for r in by_class)

    return jsonify({
        'grand_total':   grand,
        'by_class':      [{'class_id': r.cid, 'class_name': r.name, 'total': float(r.t), 'count': r.c} for r in by_class],
        'by_purpose':    [{'purpose': r.purpose, 'total': float(r.t), 'count': r.c} for r in by_purpose],
        'by_collector':  [{'name': r.full_name or r.username, 'total': float(r.t), 'count': r.c} for r in by_collector],
    })


@admin_bp.route('/reports/payments/export', methods=['GET'])
@admin_required
def export_payment_report():
    year_id = request.args.get('year_id') or _active_year_id()
    year    = AcademicYear.query.get(year_id) if year_id else None
    label   = year.label if year else 'All'

    def hrow(ws):
        hfont = Font(bold=True, color='FFFFFF')
        hfill = PatternFill(start_color='0A1628', end_color='0A1628', fill_type='solid')
        for c in ws[1]: c.font = hfont; c.fill = hfill

    wb = openpyxl.Workbook()

    ws1 = wb.active; ws1.title = 'All Payments'
    ws1.append(['Student', 'Class', 'Purpose', 'Amount (GHS)', 'Collected By', 'Confirmed', 'Date'])
    hrow(ws1)
    q = Payment.query
    if year_id: q = q.filter_by(year_id=int(year_id))
    for p in q.order_by(Payment.created_at.desc()).all():
        ws1.append([
            p.Student.name if p.student else '',
            p.student.class_.name if p.student and p.student.class_ else '',
            p.purpose, float(p.amount),
            (p.collector.full_name or p.collector.username) if p.collector else '',
            'Yes' if p.is_confirmed else 'No',
            p.created_at.strftime('%d/%m/%Y') if p.created_at else ''
        ])
    for col, w in [('A',30),('B',15),('C',28),('D',16),('E',25),('F',12),('G',14)]:
        ws1.column_dimensions[col].width = w

    ws2 = wb.create_sheet('By Class')
    ws2.append(['Class', 'Total (GHS)', 'Transactions'])
    hrow(ws2)
    by_class = (db.session.query(Class.name, func.sum(Payment.amount), func.count(Payment.id))
                .join(Student, Student.id == Payment.student_id)
                .join(Class, Class.id == Student.class_id))
    if year_id: by_class = by_class.filter(Payment.student.has(year_id=year_id) == int(year_id))
    for r in by_class.group_by(Class.name).order_by(Class.name).all():
        ws2.append([r[0], float(r[1]), r[2]])
    for col, w in [('A',24),('B',20),('C',16)]: ws2.column_dimensions[col].width = w

    ws3 = wb.create_sheet('By Purpose')
    ws3.append(['Purpose', 'Total (GHS)', 'Transactions'])
    hrow(ws3)
    by_p = db.session.query(Payment.purpose, func.sum(Payment.amount), func.count(Payment.id))
    if year_id: by_p = by_p.filter(Payment.student.has(year_id=year_id) == int(year_id))
    for r in by_p.group_by(Payment.purpose).all():
        ws3.append([r[0], float(r[1]), r[2]])
    for col, w in [('A',28),('B',20),('C',16)]: ws3.column_dimensions[col].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f'ASMaP_Payments_{label.replace("/","_")}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


# â”€â”€ MONITORING â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin_bp.route('/monitoring', methods=['GET'])
@admin_required
def list_monitoring():
    year_id    = request.args.get('year_id') or _active_year_id()
    class_id   = request.args.get('class_id')
    teacher_id = request.args.get('teacher_id')
    date_from  = request.args.get('date_from')
    date_to    = request.args.get('date_to')
    session_t  = request.args.get('session_type')

    q = MonitoringRecord.query
    if year_id:    q = q.filter_by(year_id=int(year_id))
    if class_id:   q = q.filter_by(class_id=int(class_id))
    if teacher_id: q = q.filter_by(teacher_id=int(teacher_id))
    if session_t:  q = q.filter_by(session_type=session_t)
    if date_from:  q = q.filter(MonitoringRecord.attendance_date >= date_from)
    if date_to:    q = q.filter(MonitoringRecord.attendance_date <= date_to)

    return jsonify([r.to_dict() for r in q.order_by(MonitoringRecord.attendance_date.desc()).all()])


@admin_bp.route('/monitoring/teacher-summary', methods=['GET'])
@admin_required
def teacher_summary():
    year_id = request.args.get('year_id') or _active_year_id()
    q = (db.session.query(Teacher.id, Teacher.full_name, Teacher.subject,
                          func.count(MonitoringRecord.id).label('sessions'),
                          func.sum(MonitoringRecord.students_present).label('students'))
         .join(MonitoringRecord, MonitoringRecord.teacher_id == Teacher.id))
    if year_id: q = q.filter(MonitoringRecord.id == int(year_id))
    rows = q.group_by(Teacher.id).order_by(func.count(MonitoringRecord.id).desc()).all()
    return jsonify([{'teacher_id': r.id, 'full_name': r.full_name,
                     'subject': r.subject or '', 'total_sessions': r.sessions,
                     'total_students': int(r.students or 0)} for r in rows])


@admin_bp.route('/reports/teachers', methods=['GET'])
@admin_required
def teacher_evaluation_report():
    """
    Returns per-teacher evaluation: Normal sessions vs Extended sessions,
    on-time counts, late counts, and avg students present.
    This drives the remuneration analysis.
    """
    year_id = request.args.get('year_id') or _active_year_id()
    teachers = Teacher.query.filter_by(is_active=True).order_by(Teacher.full_name).all()

    def stats(teacher_id, session_type):
        q = MonitoringRecord.query.filter_by(teacher_id=teacher_id, session_type=session_type)
        if year_id: q = q.filter_by(year_id=int(year_id))
        recs = q.all()
        total     = len(recs)
        on_time   = sum(1 for r in recs if r.is_on_time is True)
        late      = sum(1 for r in recs if r.is_on_time is False)
        unmarked  = total - on_time - late
        avg_std   = round(sum(r.students_present for r in recs) / total, 1) if total else 0
        return {'sessions': total, 'on_time': on_time, 'late': late,
                'unmarked': unmarked, 'avg_students': avg_std}

    result = []
    for t in teachers:
        normal   = stats(t.id, 'Normal')
        extended = stats(t.id, 'Extended')
        if normal['sessions'] > 0 or extended['sessions'] > 0:
            result.append({
                'teacher_id':   t.id,
                'full_name': t.full_name,
                'subject':      t.subject or '',
                'normal':       normal,
                'extended':     extended,
                'total':        normal['sessions'] + extended['sessions'],
            })

    result.sort(key=lambda x: x['total'], reverse=True)
    return jsonify(result)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# PASTE THESE ROUTES INTO YOUR EXISTING routes/admin.py
# Place them AFTER the existing /teachers GET/POST route
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# â”€â”€ GET /admin/teachers/template  â€” download blank Excel template â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin_bp.route('/teachers/template', methods=['GET'])
@jwt_required()
def teacher_template():
    claims = get_jwt()
    if claims.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from flask import send_file

    wb = Workbook()
    ws = wb.active
    ws.title = 'Teachers'

    # Header row
    headers = ['Full Name', 'Staff ID (optional)', 'Primary Subject']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1a3a5c')
        cell.alignment = Alignment(horizontal='center')

    # Example rows
    examples = [
        ['Mr. Kweku Asante', 'STF001', 'GENERAL MATHS'],
        ['Mrs. Ama Boateng', 'STF002', 'ENGLISH LANG'],
        ['Mr. Kofi Mensah',  'STF003', 'BIOLOGY'],
    ]
    for r, row in enumerate(examples, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='teachers_template.xlsx')


# â”€â”€ POST /admin/teachers/import  â€” bulk import from Excel â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin_bp.route('/teachers/import', methods=['POST'])
@jwt_required()
def import_teachers():
    claims = get_jwt()
    if claims.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Please upload an Excel file (.xlsx or .xls)'}), 400

    from openpyxl import load_workbook
    from io import BytesIO

    try:
        wb = load_workbook(BytesIO(file.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'Could not read Excel file: {str(e)}'}), 400

    created = 0
    skipped = 0
    errors  = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue  # skip blank rows

        # Read columns: Full Name | Staff ID | Primary Subject
        full_name   = str(row[0]).strip() if row[0] else ''
        staff_id    = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        subject_str = str(row[2]).strip().upper() if len(row) > 2 and row[2] else ''

        if not full_name:
            errors.append(f'Row {row_num}: Name is empty â€” skipped')
            skipped += 1
            continue

        # Check duplicate by name
        existing = Teacher.query.filter(
            db.func.lower(Teacher.full_name) == full_name.lower()
        ).first()
        if existing:
            skipped += 1
            continue

        # Match subject by name (case-insensitive)
        subject = None
        if subject_str:
            subject = Subject.query.filter(
                db.func.upper(Subject.name) == subject_str
            ).first()

        teacher = Teacher(
            full_name  = full_name,
            staff_id   = staff_id or None,
            subject_id = subject.id if subject else None,
        )
        db.session.add(teacher)
        created += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

    return jsonify({
        'message': f'Import complete: {created} added, {skipped} skipped',
        'created': created,
        'skipped': skipped,
        'errors': errors,
    })


# â”€â”€ GET /admin/teachers/export  â€” export all teachers to Excel â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin_bp.route('/teachers/export', methods=['GET'])
@jwt_required()
def export_teachers():
    claims = get_jwt()
    if claims.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from flask import send_file

    teachers = Teacher.query.order_by(Teacher.full_name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Teachers'

    headers = ['Full Name', 'Staff ID', 'Primary Subject', 'All Subjects', 'Active']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1a3a5c')

    for r, t in enumerate(teachers, 2):
        all_subjects = ', '.join(s.name for s in t.subjects) if t.subjects else (t.subject.name if t.subject else '')
        ws.cell(row=r, column=1, value=t.full_name)
        ws.cell(row=r, column=2, value=t.staff_id or '')
        ws.cell(row=r, column=3, value=t.subject.name if t.subject else '')
        ws.cell(row=r, column=4, value=all_subjects)
        ws.cell(row=r, column=5, value='Yes' if t.is_active else 'No')

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 25

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='teachers_export.xlsx')


# â”€â”€ POST /admin/classes/import  â€” bulk class creation â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin_bp.route('/classes/import', methods=['POST'])
@jwt_required()
def import_classes():
    """Accept a list of class names in JSON body: { "names": ["SHS1A", "SHS1B", ...] }"""
    claims = get_jwt()
    if claims.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403

    data = request.get_json(silent=True) or {}
    names = data.get('names', [])
    if not names:
        return jsonify({'error': 'Provide a list of class names'}), 400

    created = 0
    skipped = 0
    for name in names:
        name = str(name).strip()
        if not name:
            continue
        existing = Class.query.filter(db.func.lower(Class.name) == name.lower()).first()
        if existing:
            skipped += 1
            continue
        db.session.add(Class(name=name))
        created += 1

    db.session.commit()
    return jsonify({'message': f'{created} classes created, {skipped} already existed',
                    'created': created, 'skipped': skipped})


# â”€â”€ POST /admin/teachers/<id>/subjects  â€” assign multiple subjects to teacher â”€
@admin_bp.route('/teachers/<int:teacher_id>/subjects', methods=['POST'])
@jwt_required()
def set_teacher_subjects(teacher_id):
    claims = get_jwt()
    if claims.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403

    teacher = Teacher.query.get_or_404(teacher_id)
    data = request.get_json(silent=True) or {}
    subject_ids = data.get('subject_ids', [])

    subjects = Subject.query.filter(Subject.id.in_(subject_ids)).all()
    teacher.subjects = subjects
    db.session.commit()
    return jsonify({'message': 'Subjects updated', 'teacher': teacher.to_dict()})


# â”€â”€ GET /admin/teachers/<id>/subjects  â€” get teacher's subjects â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@admin_bp.route('/teachers/<int:teacher_id>/subjects', methods=['GET'])
@jwt_required()
def get_teacher_subjects(teacher_id):
    claims = get_jwt()
    if claims.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403

    teacher = Teacher.query.get_or_404(teacher_id)
    return jsonify({'subjects': [s.to_dict() for s in teacher.subjects]})





